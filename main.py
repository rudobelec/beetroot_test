# before running the program you need to run next commands in the terminal
# pip install openpyxl
import re
from openpyxl import load_workbook
from openpyxl.utils.exceptions import IllegalCharacterError
from openpyxl.cell.cell import ILLEGAL_CHARACTERS_RE


# function to define the first empty cell in the excel file
def first_empty():
    wb = load_workbook("task.xlsx")
    ws = wb.active
    counter = 3

    while ws[f"A{counter}"].value:
        counter += 1

    return counter


def text_to_chunks(document: str) -> list:
    with open(document, "rb") as magazine:
        text = magazine.read().decode("utf-8", errors="replace")
        split_text: list = re.split(r"(?=\bP\d{3}\b)", text)

    return split_text[1:]


# function to work with some specific data in the pdf (the one with several authors with different affiliations)
def return_digit(text: str):
    for i in text:
        if i.isdigit():
            return i

    return False


# function to clean and to normalize text. I suggest it can be improved to refine the results
def clean_text(txt: str) -> str:
    cleaned_text = txt
    patterns_to_remove = [
        r'www.medicaljournals.se/acta',
        r'\d\d?\s*5th World Psoriasis & Psoriatic Arthritis Conference 2018',
        r'POSTERS',
        r'Poster abstracts\s*\d\d?',
        r'Acta Derm Venereol 2018',
        r'\n\s*\n'
    ]
    for pattern in patterns_to_remove:
        cleaned_text = re.sub(pattern, '\n', cleaned_text)

    cleaned_text = re.sub(ILLEGAL_CHARACTERS_RE, '', cleaned_text)

    return cleaned_text


# function to create different regex pattern to find an article or to return False in case such pattern isn't found
def create_presentation_pattern(text: str):
    possible_beginnings: list = [
        "Introduction:",
        "Background:",
        "Objective:",
        "Introduction/Objective:",
        "Background/Objective:",
        "Introduction & Objectives:",
        "Introduction and Objectives:"
    ]
    beginnings: list = []

    for beginning in possible_beginnings:
        if beginning in text:
            beginnings.append(beginning)

    if not beginnings:
        return False

    return beginnings[0]


# here I decided to remove references at the end of some articles as they're a serious obstacle to regex patterns
def remove_references(text: str):
    pattern = "References:"
    if pattern in text:
        references = re.split(r"(?=References:)", text)[-1]
        references = re.escape(references)
        text = re.sub(r'{}'.format(references), '', text, flags=re.IGNORECASE)
        return text
    else:
        return text


# here I tried to create a decent function that would scrape through names
# and I'd been thinking it works great until P036
def create_name_list(text: str) -> list:
    patterns_to_avoid = [
        r'University',
        r'Clinic',
        r'Centre',
        r'Center',
        r'Department',
    ]

    text = text[0:len(text) // 4]

    session = re.findall(r"\bP\d{3}\b", text)[0]

    indexed_name_pattern = re.findall(r'([A-Z][a-z]+)(\s[A-Z][a-z]*)*(\d+)', text)

    if indexed_name_pattern:
        names = [''.join(name) for name in indexed_name_pattern]
        return names

    elements_to_remove = [session]
    lines = text.split("\n")

    for line in lines:
        if re.match(r'\b[A-Z\s]+\b', line):
            elements_to_remove.append(line)

    for element in elements_to_remove:
        text = re.sub(r'{}'.format(element), '', text)

    text = re.sub(r'\n\n', '', text)
    text = text.split('\n')

    names = []

    for line in text:
        line = line.split(', ')

        for name in line:
            name = name.strip()

            if re.match(r'([A-Z][a-z]+)(\s[A-Z]\.)*(\s[A-Z][a-z]+){1,3}', name):
                names.append(name)

    names_to_avoid = []

    for pattern in patterns_to_avoid:
        for name in names:
            if pattern in name:
                names_to_avoid.append(name)

    names = [name for name in names if name not in names_to_avoid]

    return names


# class that represents each article
class Article:
    # unfortunately I couldn't find a pattern to make affiliation and location different instance attributes
    # so here I decided to merge them
    # but I believe I can solve this later given more time
    def __init__(self, person: list, affiliation_location: list, session: str, topic_title: str, presentation: str):
        self.person = person
        self.affiliation_location = affiliation_location
        self.session = session
        self.topic_title = topic_title
        self.presentation = presentation

    # this method uploads the data from the class into the excel file in the format provided in the task
    def upload_excel(self):
        wb = load_workbook("task.xlsx")
        ws = wb.active
        counter = first_empty()

        for person in self.person:
            ws[f"A{counter + self.person.index(person)}"].value = '' .join((char for char in person if not char.isdigit()))

            for affiliation in self.affiliation_location:
                if return_digit(affiliation) == return_digit(person):
                    ws[f"B{counter + self.person.index(person)}"].value = '' .join((char for char in affiliation if not char.isdigit()))

            ws[f"D{counter + self.person.index(person)}"].value = self.session
            ws[f"E{counter + self.person.index(person)}"].value = self.topic_title
            ws[f"F{counter + self.person.index(person)}"].value = self.presentation

        wb.save("task.xlsx")


# function that creates an Article instance from a single article from the pdf
def group_article(txt: str) -> Article:
    txt = remove_references(txt)
    pattern_word = create_presentation_pattern(txt)
    lines = txt.split("\n")
    topic_title_lines = [line for line in lines[1:] if line.isupper()]
    person = create_name_list(txt)
    for element in person:
        txt = re.sub(r'{},*'.format(element), '', txt)
    session = re.findall(r"\bP\d{3}\b", txt)[0]

    if pattern_word:
        info = re.split(r'(?={})'.format(pattern_word), txt, flags=re.IGNORECASE)
        affiliation_location_lines = [item for item in info[0].split("\n") if item not in topic_title_lines][1:]
        affiliation_location = re.split(r'(?=\d[A-Z])', ''.join(affiliation_location_lines))
        affiliation_location = [item for item in affiliation_location if item != '']
        topic_title = ''.join(topic_title_lines)
        presentation = info[1]
    else:
        affiliation_location = "THE ARTICLE SEEMS TO HAVE NON-STANDARD PATTERN, YOU MIGHT NEED TO ADD IT MANUALLY"
        topic_title = ''.join(topic_title_lines)
        presentation = "THE ARTICLE SEEMS TO HAVE NON-STANDARD PATTERN, YOU MIGHT NEED TO ADD IT MANUALLY"

    return Article(person, affiliation_location, session, topic_title, presentation)


# encountered_errors list and try/except were used to avoid IllegalCharacterError previously
# here we create a list of articles and clean them so that the parsing would be more precise
def upload_all():
    split_text: list = text_to_chunks("magazine.json")

    cleaned_texts: list = [clean_text(text) for text in split_text]
    encountered_errors = []
    for text in cleaned_texts:
        try:
            article = group_article(text)
            article.upload_excel()
        except IllegalCharacterError as e:
            encountered_errors.append(article.session)
            print(f"We have encountered an Error in {article.session}")


upload_all()
# I'm aware that the results of the parsing could be better and I'm sure I can improve the results
